import openpyxl 
wb = openpyxl.load_workbook('binex.xlsx')
ws = wb.active

with open('numbers.txt') as f1:
    lines1 = f1.readlines()
r1 = 2
lines1 = [line.rstrip('\n') for line in lines1]
for line in lines1:
    ws.cell(row = r1, column = 1).value = int(line)
    r1 += 1

with open('comparisons.txt') as f2:
    lines2 = f2.readlines()
r2 = 2
lines2 = [line.rstrip('\n') for line in lines2]
for line in lines2:
    ws.cell(row = r2, column = 2).value = int(line)
    r2 += 1

wb.save('binex.xlsx')